import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import time

st.title("出荷在庫引当システム")

# ------------------------------
# Helper Functions
# ------------------------------
def sort_and_move_first(df, sort_col):
    """Sort dataframe by a column and move it to the first column."""
    df_sorted = df.sort_values(sort_col)
    cols = list(df_sorted.columns)
    cols.insert(0, cols.pop(cols.index(sort_col)))
    return df_sorted[cols]

def highlight_ordersheet(ws):
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    headers = [cell.value for cell in ws[1]]

    try:
        col_correct = headers.index("出荷数要訂正") + 1
    except ValueError:
        return

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col_correct).value
        if isinstance(val, (int, float)) and val > 0:#condition to check the order - shipping
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = highlight

def color_excel_headers(ws, color="FFD580"):
    """Apply background color to the header row of an openpyxl worksheet."""
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in ws[1]:  # first row is header
        cell.fill = header_fill
        cell.font = Font(bold=True, size=12)

# ------------------------------
# Hide columns as it reference format
# ------------------------------
COLUMNS_TO_HIDE = ["受注No", "受注行No", "出荷優先度", "倉庫CD", "倉庫名", "得意先CD", "得意先名", "ロケ", "ﾏｽﾀ単価", "金額", "現在庫数", "出荷指示数", "未出荷数", "単価相違"]

def hide_columns(ws, df, columns_to_hide):
    """Hide specific columns in the Excel sheet but leave Streamlit display unchanged."""
    for col_name in columns_to_hide:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1  # convert to 1-based index
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True

def add_group_headers(ws, group_col_name):
    headers = [cell.value for cell in ws[1]]
    try:
        col_idx = headers.index(group_col_name) + 1
    except ValueError:
        return

    ws.sheet_properties.outlinePr.summaryBelow = False

    # STEP 1 — Detect groups BEFORE inserting rows
    groups = []
    start_row = 2
    current_value = ws.cell(row=2, column=col_idx).value

    for row in range(3, ws.max_row + 2):
        value = ws.cell(row=row, column=col_idx).value
        if value != current_value:
            groups.append((current_value, start_row, row - 1))
            start_row = row
            current_value = value

    # STEP 2 — Insert headers in REVERSE so row indices remain correct
    for value, start, end in reversed(groups):
        ws.insert_rows(start)

    # STEP 3 — Recalculate exact group positions AFTER row insertion
    offset = 0
    for value, start, end in groups:
        header_row = start + offset
        first_child = header_row + 1
        last_child = end + offset + 1

        # Number of child rows
        child_count = (end - start + 1)

        # Insert header text
        ws.cell(row=header_row, column=1).value = f"{group_col_name}: {value} (合計: {child_count})"
        ws.cell(row=header_row, column=1).font = Font(name="Yu Gothlic", bold=True)
        ws.cell(row=header_row, column=1).fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")

        # Create actual row group
        ws.row_dimensions.group(first_child, last_child, outline_level=1)

        offset += 1

def set_japanese_font(ws, font_name="Yu Gothic"):
    """Apply Japanese-safe font to the entire worksheet."""
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=font_name)

def autofit_columns(ws):
    """Auto-fit column widths based on cell contents."""

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                # Larger multiplier gives more breathing room
                max_length = max(max_length, len(cell_value))
            except:
                pass

        adjusted = (max_length * 1.5 + 2)  # margin
        ws.column_dimensions[col_letter].width = adjusted

# ------------------------------
# Create downloadable excel file with 3 sheets with highlight function
# ------------------------------
def create_excel_file(sheet1, order_sheet, product_sheet):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet1.to_excel(writer, sheet_name="Sheet1", index=False)
        order_sheet.to_excel(writer, sheet_name="相手先注文", index=False)
        product_sheet.to_excel(writer, sheet_name="商品名", index=False)
        # Freeze header row on each sheet
        writer.sheets["Sheet1"].freeze_panes = "A2"
        writer.sheets["相手先注文"].freeze_panes = "B2"
        writer.sheets["商品名"].freeze_panes = "B2"

    wb = load_workbook(filename=BytesIO(output.getvalue()))

    # Setting Font Yu Gothic to all sheets:
    set_japanese_font(wb["Sheet1"], "Yu Gothic")
    set_japanese_font(wb["相手先注文"], "Yu Gothic")
    set_japanese_font(wb["商品名"], "Yu Gothic")

    # Highlighting when order > shipping
    highlight_ordersheet(wb["Sheet1"])
    highlight_ordersheet(wb["相手先注文"])
    highlight_ordersheet(wb["商品名"])
    
    # --- insert group headers ---
    add_group_headers(wb["相手先注文"], "相手先注文No")
    add_group_headers(wb["商品名"], "商品名")

    # ---- hide columns ONLY in the downloaded Excel ----
    hide_columns(wb["Sheet1"], sheet1, COLUMNS_TO_HIDE)
    hide_columns(wb["相手先注文"], order_sheet, COLUMNS_TO_HIDE)
    hide_columns(wb["商品名"], product_sheet, COLUMNS_TO_HIDE)

    # ---- Color headers in all sheets ----
    color_excel_headers(wb["Sheet1"], color="FFD580")        # Yellow headers
    color_excel_headers(wb["相手先注文"], color="FFD580")
    color_excel_headers(wb["商品名"], color="FFD580")

    autofit_columns(wb["Sheet1"])
    autofit_columns(wb["相手先注文"])
    autofit_columns(wb["商品名"])

    out2 = BytesIO()

    wb.save(out2)
    return out2.getvalue()

# ------------------------------
# Create two separate excel files based on product id
# ------------------------------
def filter_by_product_id(df, prefix="15"):
    df = df.copy()
    df["商品CD"] = df["商品CD"].astype(str)
    return df[df["商品CD"].str.startswith(prefix)], df[~df["商品CD"].str.startswith(prefix)]

# ------------------------------
# Upload CSV
# ------------------------------
uploaded_file = st.file_uploader("アップロード CSV", type=["csv"])

if uploaded_file:
    df = pd.read_csv(uploaded_file, header=1, encoding="cp932")  # second row as header

    # Convert numeric columns' NaN/empty to zero
    numeric_cols = df.select_dtypes(include=["number"]).columns
    df[numeric_cols] = df[numeric_cols].fillna(0)
    
    # Create a placeholder
    placeholder = st.empty()
    progress = st.progress(0, text="0%")
    # Drop last 2 columns if possible
    if df.shape[1] > 2:
        df = df.drop(columns=[df.columns[-2], df.columns[-1]])

    # ------------------------------
    # Calculated Columns
    # ------------------------------
    required_cols = ["単価", "出荷数", "受注数"]
    if all(col in df.columns for col in required_cols):
        df["出荷金額"] = df["単価"] * df["出荷数"]
        df["出荷数要訂正"] = df["受注数"] - df["出荷数"]
        df["受注金額"] = df["受注数"] * df["単価"]
        df["欠品金額"] = df["出荷数要訂正"] * df["単価"]
        # Show the temporary message
        placeholder.success("エクセルの準備...")
    else:
        st.error(f"Missing required columns: {required_cols}")

    # ------------------------------
    # Prepare sheets
    # ------------------------------
    order_sheet = sort_and_move_first(df, "相手先注文No") if "相手先注文No" in df.columns else pd.DataFrame()
    product_sheet = sort_and_move_first(df, "商品名") if "商品名" in df.columns else pd.DataFrame()

    # Full Excel file
    full_excel = create_excel_file(df, order_sheet, product_sheet)
    st.subheader("ダウンロードエクセル")
    st.download_button(
        "すべてダウンロード",
        full_excel,
        file_name="出荷在庫引当.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    progress.progress(33, text="33%")

    # ------------------------------
    # Split files by 商品CD prefix "15"
    # ------------------------------
    if "商品CD" in df.columns:
        dfA, dfB = filter_by_product_id(df, "15")

        # File A
        order_sheet_A = sort_and_move_first(dfA, "相手先注文No")
        product_sheet_A = sort_and_move_first(dfA, "商品名")
        excelA = create_excel_file(dfA, order_sheet_A, product_sheet_A)
        st.download_button(
            "ダウンロード File A (商品CD 15-)",
            excelA,
            file_name="出荷在庫引当_4251.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        progress.progress(66, text="66%")

        # File B
        order_sheet_B = sort_and_move_first(dfB, "相手先注文No")
        product_sheet_B = sort_and_move_first(dfB, "商品名")
        excelB = create_excel_file(dfB, order_sheet_B, product_sheet_B)
        st.download_button(
            "ダウンロード File B (商品CD 15- 以外)",
            excelB,
            file_name="出荷在庫引当_9052.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        progress.progress(100, text="100%")
        time.sleep(2)
        progress.empty()
        # Remove the message
        placeholder.empty()
    else:
        st.error("Column '商品CD' not found — cannot split the file.")



















